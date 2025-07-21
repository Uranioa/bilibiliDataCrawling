import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from module.Format import *


def is_excel_file ( file_path: str ) -> bool:
	file_path = path_formatting ( file_path )
	try:
		openpyxl.load_workbook ( file_path )
		return True
	except (InvalidFileException, FileNotFoundError):
		return False


def read_excel ( file_path ):
	file_path = path_formatting ( file_path )
	workbook = openpyxl.load_workbook ( file_path )
	sheet = workbook.active
	return [ row [ 0 ].value for row in sheet.iter_rows ( min_row = 2, min_col = 2, max_col = 2 ) ]


def write_to_excel ( file_path, idx, data ):
	file_path = path_formatting ( file_path )
	workbook = openpyxl.load_workbook ( file_path )
	sheet = workbook.active
	sheet.cell ( row = idx + 1, column = 1, value = data [ "video_title" ] )
	sheet.cell ( row = idx + 1, column = 3, value = data [ "release_time" ] )
	sheet.cell ( row = idx + 1, column = 4, value = data [ "video_playback" ] )
	sheet.cell ( row = idx + 1, column = 5, value = data [ "video_barrage" ] )
	sheet.cell ( row = idx + 1, column = 6, value = data [ "video_likes" ] )
	sheet.cell ( row = idx + 1, column = 7, value = data [ "video_coin_drop" ] )
	sheet.cell ( row = idx + 1, column = 8, value = data [ "video_collection" ] )
	sheet.cell ( row = idx + 1, column = 9, value = data [ "video_sharing" ] )
	workbook.save ( file_path )
	print ( f"视频：{data [ "video_title" ]}，数据爬取完成，已写入 Excel 文件。" )
